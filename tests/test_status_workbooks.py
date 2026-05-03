from __future__ import annotations

from pathlib import Path

from langchain_core.documents import Document
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from agentic_chatbot_next.rag.ingest import _build_chunk_records
from agentic_chatbot_next.rag.status_workbooks import (
    classify_headers,
    extract_actions,
    extract_budgets,
    extract_cdrls,
    extract_issues,
    extract_milestones,
    extract_requirements,
    extract_risks,
    extract_schedules,
    extract_status_records,
    extract_test_events,
    profile_workbook,
)
from agentic_chatbot_next.rag.tabular import (
    deterministic_status_evidence_results,
    plan_tabular_evidence_tasks,
    tabular_evidence_results_to_documents,
)
from agentic_chatbot_next.rag.workbook_loader import load_workbook_documents


def _make_status_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Status"
    sheet.append(["Risk ID", "Risk Description", "Owner", "Status", "Due Date", "Mitigation"])
    sheet.append(["R-1", "Battery supplier late", "Pat Lee", "Open", "2028-09-26", "Expedite cells"])
    table = Table(displayName="RiskTable", ref="A1:F2")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    sheet.add_table(table)
    sheet.merge_cells("H1:I1")

    actions = workbook.create_sheet("Actions")
    actions.append(["Action ID", "Action", "Owner", "Priority", "Status", "Finish Date"])
    actions.append(["A-12", "Close supplier recovery plan", "Sam Rivera", "High", "In Progress", "2028-10-01"])

    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    hidden.append(["Issue ID", "Issue", "Owner", "Status"])
    hidden.append(["I-9", "Hidden issue", "No One", "Open"])
    workbook.save(path)


def test_profile_workbook_detects_status_headers_tables_hidden_sheets_and_merged_cells(tmp_path: Path) -> None:
    path = tmp_path / "tracker.xlsx"
    _make_status_workbook(path)

    profile = profile_workbook(path)

    assert profile.workbook_name == "tracker.xlsx"
    status_sheet = next(sheet for sheet in profile.sheets if sheet.sheet_name == "Status")
    hidden_sheet = next(sheet for sheet in profile.sheets if sheet.sheet_name == "Hidden")
    assert status_sheet.header_row == 1
    assert status_sheet.named_tables[0]["name"] == "RiskTable"
    assert "H1:I1" in status_sheet.merged_cells
    assert "risk" in status_sheet.domain_tags
    assert "status" in status_sheet.domain_tags
    assert hidden_sheet.visible is False


def test_classify_headers_maps_owner_status_dates_and_budget_roles() -> None:
    roles = classify_headers(["WBS", "Owner", "Current Status", "Forecast Finish", "Budget Variance"])
    pairs = {(role.header, role.role) for role in roles}

    assert ("Owner", "owner") in pairs
    assert ("Current Status", "status") in pairs
    assert ("Forecast Finish", "finish_date") in pairs
    assert ("Budget Variance", "budget_amount") in pairs
    assert any(role.domain == "budget" for role in roles)


def test_extract_status_records_preserves_sheet_row_and_cell_range(tmp_path: Path) -> None:
    path = tmp_path / "tracker.xlsx"
    _make_status_workbook(path)

    records = extract_status_records(path, domains=["risk"], doc_id="doc-risk", title="tracker.xlsx")

    assert records
    first = records[0]
    assert first.domain == "risk"
    assert first.owner == "Pat Lee"
    assert first.status == "Open"
    assert first.source_ref.doc_id == "doc-risk"
    assert first.source_ref.sheet_name == "Status"
    assert first.source_ref.row_start == 2
    assert first.source_ref.cell_range == "A2:F2"


def test_domain_specific_extractors_cover_status_workbook_domains(tmp_path: Path) -> None:
    path = tmp_path / "domains.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "All"
    sheet.append(["ID", "Type", "Description", "Owner", "Status", "Due Date", "Budget Amount"])
    rows = [
        ["R-1", "Risk", "Supplier risk", "A", "Open", "2028-01-01", ""],
        ["I-1", "Issue", "Test issue", "B", "Open", "2028-01-02", ""],
        ["A-1", "Action", "Follow-up action", "C", "Open", "2028-01-03", ""],
        ["S-1", "Schedule", "Forecast finish", "D", "Late", "2028-01-04", ""],
        ["B-1", "Budget", "Cost variance", "E", "Watch", "2028-01-05", "$5M"],
        ["C-1", "CDRL", "CDRL A001 delivery", "F", "Due", "2028-01-06", ""],
        ["REQ-1", "Requirement", "Shall verify telemetry", "G", "Compliant", "2028-01-07", ""],
        ["T-1", "Test Event", "TRR readiness", "H", "Planned", "2028-01-08", ""],
        ["M-1", "Milestone", "CDR review", "I", "Approved", "2028-01-09", ""],
    ]
    for row in rows:
        sheet.append(row)
    workbook.save(path)

    extractors = [
        extract_risks,
        extract_issues,
        extract_actions,
        extract_schedules,
        extract_budgets,
        extract_cdrls,
        extract_requirements,
        extract_test_events,
        extract_milestones,
    ]

    assert all(extractor(path) for extractor in extractors)


def test_workbook_loader_enriches_row_chunks_with_status_metadata(tmp_path: Path) -> None:
    path = tmp_path / "tracker.xlsx"
    _make_status_workbook(path)

    docs = load_workbook_documents(path)
    row = next(doc for doc in docs if doc.metadata.get("sheet_name") == "Status" and doc.metadata.get("row_start") == 2)

    assert "risk" in row.metadata["status_domains"]
    assert row.metadata["status_workbook"]["key_values"]["owner"] == "Pat Lee"
    assert "Status Domains: risk" in row.page_content


def test_chunk_records_preserve_status_metadata_json(tmp_path: Path) -> None:
    path = tmp_path / "tracker.xlsx"
    _make_status_workbook(path)
    row_doc = next(
        doc
        for doc in load_workbook_documents(path)
        if doc.metadata.get("sheet_name") == "Status" and doc.metadata.get("row_start") == 2
    )

    records = _build_chunk_records([row_doc], "doc-status", collection_id="default")

    assert records[0].metadata_json["status_domains"] == ["risk", "schedule", "status"]
    assert records[0].metadata_json["status_workbook"]["key_values"]["status"] == "Open"


def test_deterministic_status_evidence_results_create_citable_tabular_docs(tmp_path: Path) -> None:
    path = tmp_path / "tracker.xlsx"
    _make_status_workbook(path)
    planning_doc = Document(
        page_content="Workbook: tracker.xlsx | Sheet: Status | Row 2: Risk Description: Battery supplier late",
        metadata={
            "doc_id": "doc-status",
            "title": "tracker.xlsx",
            "file_type": "xlsx",
            "source_path": str(path),
            "sheet_name": "Status",
            "row_start": 2,
            "row_end": 2,
            "cell_range": "A2:F2",
        },
    )

    tasks = plan_tabular_evidence_tasks("List the open risks with owner and due date.", [planning_doc])
    results, warnings, stats = deterministic_status_evidence_results("List the open risks with owner and due date.", tasks)
    docs = tabular_evidence_results_to_documents(results, tasks)

    assert warnings == []
    assert stats["record_count"] >= 1
    assert docs
    assert docs[0].metadata["sheet_name"] == "Status"
    assert docs[0].metadata["row_start"] == 2
    assert docs[0].metadata["cell_range"] == "A2:F2"
    assert "Battery supplier late" in docs[0].page_content

