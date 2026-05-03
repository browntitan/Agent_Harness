"""Unit tests for data analyst tools (make_data_analyst_tools).

All external dependencies (stores, DockerSandboxExecutor) are mocked.
Real pandas operations are used for load_dataset and inspect_columns tests
so that the stat computation logic is validated end-to-end.
"""
from __future__ import annotations

import json
import tempfile
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

import pytest


# ---------------------------------------------------------------------------
# Fixtures and helpers
# ---------------------------------------------------------------------------

def _make_session(scratchpad=None):
    session = MagicMock()
    session.scratchpad = scratchpad or {}
    session.tenant_id = "test-tenant"
    session.metadata = {}
    session.workspace = None
    return session


def _make_stores(source_path=None):
    stores = MagicMock()
    doc = MagicMock()
    doc.source_path = source_path
    doc.source_uri = f"file://{source_path}" if source_path else None
    doc.collection_id = "default"
    stores.doc_store.get_document.return_value = doc
    return stores


def _make_settings():
    settings = MagicMock()
    settings.sandbox_docker_image = "agentic-chatbot-sandbox:py312"
    settings.sandbox_timeout_seconds = 30
    settings.sandbox_memory_limit = "256m"
    settings.data_analyst_nlp_chat_model = ""
    settings.data_analyst_nlp_batch_size = 5
    settings.data_analyst_nlp_temperature = 0.0
    settings.ollama_chat_model = "base-chat"
    settings.azure_openai_chat_deployment = ""
    settings.nvidia_chat_model = ""
    return settings


def _write_csv(path: Path, content: str) -> None:
    path.write_text(content)


def _write_xlsx(path: Path) -> None:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["name", "value", "category"])
    ws.append(["Alice", 100, "A"])
    ws.append(["Bob", 200, "B"])
    ws.append(["Carol", 150, "A"])
    wb.save(str(path))


def _write_multisheet_xlsx(path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "raw_reviews"
    ws.append(["review", "rating"])
    ws.append(["Loved it", 5])
    ws.append(["Could be better", 3])

    meta = wb.create_sheet("metadata")
    meta.append(["key", "value"])
    meta.append(["source", "survey"])
    wb.save(str(path))


def _write_status_xlsx(path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Risks"
    ws.append(["Risk ID", "Risk", "Owner", "Status", "Due Date"])
    ws.append(["R-1", "Supplier delay", "Pat Lee", "Open", "2028-09-26"])
    wb.save(str(path))


class _RecordingChatModel:
    def __init__(self, responses):
        self.responses = list(responses)
        self.calls = []

    def invoke(self, messages, config=None):
        self.calls.append({"messages": list(messages), "config": dict(config or {})})
        if not self.responses:
            raise AssertionError("No fake NLP responses remaining")
        return SimpleNamespace(content=self.responses.pop(0))


# ---------------------------------------------------------------------------
# load_dataset — CSV
# ---------------------------------------------------------------------------

class TestLoadDatasetCsv:
    def test_returns_expected_keys(self, tmp_path):
        csv_file = tmp_path / "data.csv"
        _write_csv(csv_file, "a,b,c\n1,2,3\n4,5,6\n7,8,9\n")

        session = _make_session()
        stores = _make_stores(source_path=str(csv_file))
        settings = _make_settings()

        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools
        tools = make_data_analyst_tools(stores, session, settings=settings)
        load_tool = next(t for t in tools if t.name == "load_dataset")

        result = json.loads(load_tool.invoke({"doc_id": "doc123"}))

        assert "file_path" in result
        assert "doc_id" in result
        assert "columns" in result
        assert "shape" in result
        assert "dtypes" in result
        assert "head" in result
        assert "info_summary" in result

    def test_shape_correct(self, tmp_path):
        csv_file = tmp_path / "data.csv"
        _write_csv(csv_file, "a,b\n1,2\n3,4\n5,6\n")

        session = _make_session()
        stores = _make_stores(source_path=str(csv_file))
        settings = _make_settings()

        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools
        tools = make_data_analyst_tools(stores, session, settings=settings)
        load_tool = next(t for t in tools if t.name == "load_dataset")

        result = json.loads(load_tool.invoke({"doc_id": "doc123"}))

        assert result["shape"] == [3, 2]
        assert result["columns"] == ["a", "b"]

    def test_scratchpad_populated(self, tmp_path):
        csv_file = tmp_path / "data.csv"
        _write_csv(csv_file, "x,y\n1,2\n")

        session = _make_session()
        stores = _make_stores(source_path=str(csv_file))
        settings = _make_settings()

        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools
        tools = make_data_analyst_tools(stores, session, settings=settings)
        load_tool = next(t for t in tools if t.name == "load_dataset")
        load_tool.invoke({"doc_id": "doc_abc"})

        assert "dataset_doc_abc" in session.scratchpad
        assert session.scratchpad["dataset_doc_abc"] == str(csv_file)

    def test_head_has_at_most_5_rows(self, tmp_path):
        csv_file = tmp_path / "data.csv"
        rows = "\n".join([f"{i},{i*2}" for i in range(20)])
        _write_csv(csv_file, f"a,b\n{rows}")

        session = _make_session()
        stores = _make_stores(source_path=str(csv_file))
        settings = _make_settings()

        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools
        tools = make_data_analyst_tools(stores, session, settings=settings)
        load_tool = next(t for t in tools if t.name == "load_dataset")
        result = json.loads(load_tool.invoke({"doc_id": "doc123"}))

        assert len(result["head"]) <= 5


# ---------------------------------------------------------------------------
# load_dataset — Excel
# ---------------------------------------------------------------------------

class TestLoadDatasetXlsx:
    def test_loads_excel_file(self, tmp_path):
        pytest.importorskip("openpyxl")
        xlsx_file = tmp_path / "data.xlsx"
        _write_xlsx(xlsx_file)

        session = _make_session()
        stores = _make_stores(source_path=str(xlsx_file))
        settings = _make_settings()

        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools
        tools = make_data_analyst_tools(stores, session, settings=settings)
        load_tool = next(t for t in tools if t.name == "load_dataset")
        result = json.loads(load_tool.invoke({"doc_id": "xlsdoc"}))

        assert "error" not in result
        assert result["columns"] == ["name", "value", "category"]
        assert result["shape"][0] == 3  # 3 data rows

    def test_profile_dataset_profiles_all_workbook_sheets(self, tmp_path):
        pytest.importorskip("openpyxl")
        xlsx_file = tmp_path / "reviews.xlsx"
        _write_multisheet_xlsx(xlsx_file)

        session = _make_session()
        stores = _make_stores(source_path=str(xlsx_file))
        settings = _make_settings()

        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools
        tools = make_data_analyst_tools(stores, session, settings=settings)
        profile_tool = next(t for t in tools if t.name == "profile_dataset")
        result = json.loads(profile_tool.invoke({"doc_id": "reviews_doc"}))

        assert result["status"] == "ok"
        assert result["sheet_names"] == ["raw_reviews", "metadata"]
        assert [sheet["sheet_name"] for sheet in result["sheets"]] == ["raw_reviews", "metadata"]
        assert result["source_refs"][0]["sheet_name"] == "raw_reviews"
        assert result["source_refs"][0]["cell_range"]
        assert "review" in result["sheets"][0]["columns"]

    def test_status_workbook_tools_profile_and_extract_records(self, tmp_path):
        pytest.importorskip("openpyxl")
        xlsx_file = tmp_path / "status.xlsx"
        _write_status_xlsx(xlsx_file)

        session = _make_session()
        stores = _make_stores(source_path=str(xlsx_file))
        settings = _make_settings()

        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools

        tools = make_data_analyst_tools(stores, session, settings=settings)
        profile_tool = next(t for t in tools if t.name == "profile_workbook_status")
        extract_tool = next(t for t in tools if t.name == "extract_workbook_status")

        profile = json.loads(profile_tool.invoke({"doc_id": "status_doc"}))
        extracted = json.loads(extract_tool.invoke({"doc_id": "status_doc", "domains_csv": "risk"}))

        assert profile["status"] == "ok"
        assert "risk" in profile["sheets"][0]["domain_tags"]
        assert extracted["record_count"] == 1
        assert extracted["source_refs"][0]["cell_range"] == "A2:E2"
        assert "HUMAN_REVIEW_REQUIRED_BEFORE_EXTERNAL_SHARING" in extracted["warnings"]

    def test_status_workbook_tools_respect_collection_access_summary(self, tmp_path):
        pytest.importorskip("openpyxl")
        xlsx_file = tmp_path / "status.xlsx"
        _write_status_xlsx(xlsx_file)

        session = _make_session()
        session.metadata = {
            "access_summary": {
                "authz_enabled": True,
                "session_upload_collection_id": "",
                "resources": {
                    "collection": {
                        "use": ["other-collection"],
                        "use_all": False,
                    }
                },
            }
        }
        stores = _make_stores(source_path=str(xlsx_file))
        settings = _make_settings()

        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools

        tools = make_data_analyst_tools(stores, session, settings=settings)
        profile_tool = next(t for t in tools if t.name == "profile_workbook_status")
        result = json.loads(profile_tool.invoke({"doc_id": "status_doc"}))

        assert result["status"] == "error"
        assert "Access denied" in result["error"]


# ---------------------------------------------------------------------------
# load_dataset — error cases
# ---------------------------------------------------------------------------

class TestLoadDatasetErrors:
    def test_invalid_extension_returns_error(self, tmp_path):
        txt_file = tmp_path / "readme.txt"
        txt_file.write_text("just text")

        session = _make_session()
        stores = _make_stores(source_path=str(txt_file))
        settings = _make_settings()

        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools
        tools = make_data_analyst_tools(stores, session, settings=settings)
        load_tool = next(t for t in tools if t.name == "load_dataset")
        result = json.loads(load_tool.invoke({"doc_id": "bad_doc"}))

        assert "error" in result
        assert "Unsupported" in result["error"]

    def test_missing_document_returns_error(self):
        session = _make_session()
        stores = MagicMock()
        stores.doc_store.get_document.return_value = None
        settings = _make_settings()

        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools
        tools = make_data_analyst_tools(stores, session, settings=settings)
        load_tool = next(t for t in tools if t.name == "load_dataset")
        result = json.loads(load_tool.invoke({"doc_id": "nonexistent"}))

        assert "error" in result
        assert "not found" in result["error"].lower()

    def test_file_not_on_disk_returns_error(self, tmp_path):
        session = _make_session()
        stores = _make_stores(source_path=str(tmp_path / "ghost.csv"))
        settings = _make_settings()

        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools
        tools = make_data_analyst_tools(stores, session, settings=settings)
        load_tool = next(t for t in tools if t.name == "load_dataset")
        result = json.loads(load_tool.invoke({"doc_id": "ghost_doc"}))

        assert "error" in result

    def test_next_runtime_load_dataset_can_resolve_workspace_file(self, tmp_path):
        csv_file = tmp_path / "workspace_data.csv"
        _write_csv(csv_file, "region,spend\nNA,10\nEU,20\n")

        session = _make_session()
        session.workspace = MagicMock()
        session.workspace.root = tmp_path
        session.workspace.exists.side_effect = lambda filename: (tmp_path / filename).exists()
        stores = MagicMock()
        stores.doc_store.get_document.return_value = None
        settings = _make_settings()

        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools

        tools = make_data_analyst_tools(stores, session, settings=settings)
        load_tool = next(t for t in tools if t.name == "load_dataset")
        result = json.loads(load_tool.invoke({"doc_id": "workspace_data.csv"}))

        assert "error" not in result
        assert result["doc_id"] == "workspace_data.csv"
        assert result["shape"] == [2, 2]

    def test_manifest_present_text_upload_is_unsupported_not_missing(self, tmp_path):
        txt_file = tmp_path / "readme.txt"
        txt_file.write_text("notes", encoding="utf-8")

        session = _make_session()
        session.metadata = {
            "last_upload_manifest": {
                "workspace_copies": ["readme.txt"],
                "filenames": ["readme.txt"],
            }
        }
        session.workspace = MagicMock()
        session.workspace.root = tmp_path
        session.workspace.exists.side_effect = lambda filename: (tmp_path / filename).exists()
        stores = MagicMock()
        stores.doc_store.get_document.return_value = None
        settings = _make_settings()

        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools

        tools = make_data_analyst_tools(stores, session, settings=settings)
        load_tool = next(t for t in tools if t.name == "load_dataset")
        result = json.loads(load_tool.invoke({"doc_id": "readme.txt"}))

        assert "error" in result
        assert "unsupported" in result["error"].lower()
        assert "not found" not in result["error"].lower()

    def test_next_runtime_load_dataset_defaults_to_first_workspace_file_when_doc_id_missing(self, tmp_path):
        csv_file = tmp_path / "workspace_data.csv"
        _write_csv(csv_file, "region,spend\nNA,10\nEU,20\n")

        session = _make_session()
        session.workspace = MagicMock()
        session.workspace.root = tmp_path
        session.workspace.exists.side_effect = lambda filename: (tmp_path / filename).exists()
        session.workspace.list_files.return_value = ["workspace_data.csv"]
        stores = MagicMock()
        stores.doc_store.get_document.return_value = None
        settings = _make_settings()

        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools

        tools = make_data_analyst_tools(stores, session, settings=settings)
        load_tool = next(t for t in tools if t.name == "load_dataset")
        result = json.loads(load_tool.invoke({}))

        assert "error" not in result
        assert result["doc_id"] == "workspace_data.csv"


# ---------------------------------------------------------------------------
# inspect_columns — numeric
# ---------------------------------------------------------------------------

class TestInspectColumnsNumeric:
    def test_numeric_columns_have_stats(self, tmp_path):
        csv_file = tmp_path / "nums.csv"
        _write_csv(csv_file, "score,value\n10,100\n20,200\n30,300\n40,400\n")

        session = _make_session()
        session.scratchpad["dataset_d1"] = str(csv_file)
        session.scratchpad["dataset_d1_ext"] = ".csv"
        stores = _make_stores()
        settings = _make_settings()

        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools
        tools = make_data_analyst_tools(stores, session, settings=settings)
        inspect_tool = next(t for t in tools if t.name == "inspect_columns")
        result = json.loads(inspect_tool.invoke({"doc_id": "d1", "columns": "score"}))

        assert "score" in result
        col = result["score"]
        assert "mean" in col
        assert "std" in col
        assert "min" in col
        assert "max" in col
        assert "nulls" in col

    def test_correct_mean_computed(self, tmp_path):
        csv_file = tmp_path / "nums.csv"
        _write_csv(csv_file, "v\n10\n20\n30\n")

        session = _make_session()
        session.scratchpad["dataset_d2"] = str(csv_file)
        session.scratchpad["dataset_d2_ext"] = ".csv"
        stores = _make_stores()
        settings = _make_settings()

        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools
        tools = make_data_analyst_tools(stores, session, settings=settings)
        inspect_tool = next(t for t in tools if t.name == "inspect_columns")
        result = json.loads(inspect_tool.invoke({"doc_id": "d2", "columns": "v"}))

        assert abs(result["v"]["mean"] - 20.0) < 0.001


# ---------------------------------------------------------------------------
# inspect_columns — string
# ---------------------------------------------------------------------------

class TestInspectColumnsString:
    def test_string_columns_have_top_values(self, tmp_path):
        csv_file = tmp_path / "cats.csv"
        _write_csv(csv_file, "cat\nA\nA\nB\nC\nA\n")

        session = _make_session()
        session.scratchpad["dataset_d3"] = str(csv_file)
        session.scratchpad["dataset_d3_ext"] = ".csv"
        stores = _make_stores()
        settings = _make_settings()

        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools
        tools = make_data_analyst_tools(stores, session, settings=settings)
        inspect_tool = next(t for t in tools if t.name == "inspect_columns")
        result = json.loads(inspect_tool.invoke({"doc_id": "d3", "columns": "cat"}))

        assert "cat" in result
        col = result["cat"]
        assert "top_values" in col
        assert "nulls" in col
        assert "unique" in col


# ---------------------------------------------------------------------------
# inspect_columns — before load_dataset
# ---------------------------------------------------------------------------

class TestInspectColumnsBeforeLoad:
    def test_returns_error_when_not_loaded(self):
        session = _make_session()  # empty scratchpad
        stores = _make_stores()
        settings = _make_settings()

        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools
        tools = make_data_analyst_tools(stores, session, settings=settings)
        inspect_tool = next(t for t in tools if t.name == "inspect_columns")
        result = json.loads(inspect_tool.invoke({"doc_id": "not_loaded", "columns": ""}))

        assert "error" in result
        assert "load_dataset" in result["error"]


# ---------------------------------------------------------------------------
# execute_code
# ---------------------------------------------------------------------------

class TestExecuteCodeSuccess:
    def test_returns_stdout_on_success(self, tmp_path):
        from agentic_chatbot_next.sandbox.docker_exec import SandboxResult

        csv_file = tmp_path / "data.csv"
        _write_csv(csv_file, "a,b\n1,2\n")
        session = _make_session()
        session.scratchpad["dataset_d_exec"] = str(csv_file)
        stores = _make_stores()
        settings = _make_settings()

        mock_result = SandboxResult(
            stdout="42", stderr="", exit_code=0, execution_time_seconds=0.5
        )

        with patch(
            "agentic_chatbot_next.tools.data_analyst_tools.DockerSandboxExecutor"
        ) as MockExec:
            MockExec.return_value.execute.return_value = mock_result

            from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools
            tools = make_data_analyst_tools(stores, session, settings=settings)
            exec_tool = next(t for t in tools if t.name == "execute_code")
            result = json.loads(exec_tool.invoke({
                "code": "print(42)",
                "doc_ids": "d_exec",
            }))

        assert result["stdout"] == "42"
        assert result["success"] is True


class TestExecuteCodeError:
    def test_stderr_returned_on_failure(self, tmp_path):
        from agentic_chatbot_next.sandbox.docker_exec import SandboxResult

        csv_file = tmp_path / "data.csv"
        _write_csv(csv_file, "a\n1\n")
        session = _make_session()
        session.scratchpad["dataset_d_err"] = str(csv_file)
        stores = _make_stores()
        settings = _make_settings()

        mock_result = SandboxResult(
            stdout="",
            stderr="NameError: name 'x' is not defined",
            exit_code=1,
            execution_time_seconds=0.1,
        )

        with patch(
            "agentic_chatbot_next.tools.data_analyst_tools.DockerSandboxExecutor"
        ) as MockExec:
            MockExec.return_value.execute.return_value = mock_result

            from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools
            tools = make_data_analyst_tools(stores, session, settings=settings)
            exec_tool = next(t for t in tools if t.name == "execute_code")
            result = json.loads(exec_tool.invoke({
                "code": "print(x)",
                "doc_ids": "d_err",
            }))

        assert result["success"] is False
        assert "NameError" in result["stderr"]


class TestExecuteCodeTimeout:
    def test_timeout_message_returned(self):
        from agentic_chatbot_next.sandbox.docker_exec import SandboxResult

        session = _make_session()
        stores = _make_stores()
        settings = _make_settings()

        timeout_result = SandboxResult(
            stdout="",
            stderr="Execution timed out after 30s.",
            exit_code=-1,
            execution_time_seconds=30.0,
        )

        with patch(
            "agentic_chatbot_next.tools.data_analyst_tools.DockerSandboxExecutor"
        ) as MockExec:
            MockExec.return_value.execute.return_value = timeout_result

            from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools
            tools = make_data_analyst_tools(stores, session, settings=settings)
            exec_tool = next(t for t in tools if t.name == "execute_code")
            result = json.loads(exec_tool.invoke({"code": "import time; time.sleep(999)"}))

        assert result["success"] is False
        assert "timed out" in result["stderr"].lower()


class TestExecuteCodeDockerUnavailable:
    def test_graceful_error_when_docker_unavailable(self):
        from agentic_chatbot_next.sandbox.exceptions import SandboxUnavailableError

        session = _make_session()
        stores = _make_stores()
        settings = _make_settings()

        with patch(
            "agentic_chatbot_next.tools.data_analyst_tools.DockerSandboxExecutor"
        ) as MockExec:
            MockExec.return_value.execute.side_effect = SandboxUnavailableError("Docker not running")

            from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools
            tools = make_data_analyst_tools(stores, session, settings=settings)
            exec_tool = next(t for t in tools if t.name == "execute_code")
            result = json.loads(exec_tool.invoke({"code": "print('hello')"}))

        assert "error" in result
        assert result["success"] is False


# ---------------------------------------------------------------------------
# Scratchpad tools
# ---------------------------------------------------------------------------

class TestScratchpadTools:
    def _get_tools(self):
        session = _make_session()
        stores = _make_stores()
        settings = _make_settings()

        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools
        tools = make_data_analyst_tools(stores, session, settings=settings)
        write = next(t for t in tools if t.name == "scratchpad_write")
        read = next(t for t in tools if t.name == "scratchpad_read")
        lst = next(t for t in tools if t.name == "scratchpad_list")
        return write, read, lst, session

    def test_write_then_read(self):
        write, read, _, session = self._get_tools()
        write.invoke({"key": "plan", "value": "step 1: inspect"})
        result = json.loads(read.invoke({"key": "plan"}))
        assert result["value"] == "step 1: inspect"

    def test_read_missing_key_returns_available(self):
        _, read, _, session = self._get_tools()
        result = json.loads(read.invoke({"key": "nonexistent"}))
        assert "error" in result
        assert "available_keys" in result

    def test_list_returns_user_keys(self):
        write, _, lst, session = self._get_tools()
        write.invoke({"key": "data_overview", "value": "3 rows x 5 cols"})
        write.invoke({"key": "analysis_plan", "value": "group by region"})
        result = json.loads(lst.invoke({}))
        assert "data_overview" in result["keys"]
        assert "analysis_plan" in result["keys"]
        assert result["count"] >= 2


class TestMultiSheetDatasetSupport:
    def test_load_dataset_returns_sheet_inventory_for_excel(self, tmp_path):
        pytest.importorskip("openpyxl")
        xlsx_file = tmp_path / "reviews.xlsx"
        _write_multisheet_xlsx(xlsx_file)

        session = _make_session()
        stores = _make_stores(source_path=str(xlsx_file))
        settings = _make_settings()

        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools

        tools = make_data_analyst_tools(stores, session, settings=settings)
        load_tool = next(t for t in tools if t.name == "load_dataset")
        result = json.loads(load_tool.invoke({"doc_id": "reviews_doc", "sheet_name": "metadata"}))

        assert result["sheet_name"] == "metadata"
        assert result["sheet_names"] == ["raw_reviews", "metadata"]
        assert result["columns"] == ["key", "value"]


class TestRunNlpColumnTask:
    def _workspace_session(self, tmp_path: Path):
        from agentic_chatbot_next.sandbox.workspace import SessionWorkspace

        session = _make_session()
        workspace = SessionWorkspace.for_session("tenant:user:conv", tmp_path / "workspaces")
        workspace.open()
        session.workspace = workspace
        session.session_id = "tenant:user:conv"
        session.conversation_id = "conv"
        return session

    def test_batches_rows_and_dedupes_identical_texts(self, tmp_path):
        csv_file = tmp_path / "reviews.csv"
        _write_csv(
            csv_file,
            "reviews\nAmazing product\nTerrible support\nOkay overall\nAmazing product\nNot bad\nLoved the speed\nConfusing docs\n",
        )

        session = self._workspace_session(tmp_path)
        stores = _make_stores(source_path=str(csv_file))
        settings = _make_settings()
        chat = _RecordingChatModel(
            [
                json.dumps(
                    {
                        "results": [
                            {"item_id": "row_1", "label": "positive", "score": 0.99},
                            {"item_id": "row_2", "label": "negative", "score": 0.96},
                            {"item_id": "row_3", "label": "neutral", "score": 0.55},
                            {"item_id": "row_4", "label": "positive", "score": 0.88},
                            {"item_id": "row_5", "label": "positive", "score": 0.91},
                        ]
                    }
                ),
                json.dumps(
                    {
                        "results": [
                            {"item_id": "row_6", "label": "negative", "score": 0.79},
                        ]
                    }
                ),
            ]
        )

        from agentic_chatbot_next.providers import ProviderBundle
        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools

        tools = make_data_analyst_tools(
            stores,
            session,
            settings=settings,
            providers=ProviderBundle(chat=chat, judge=chat, embeddings=object()),
        )
        nlp_tool = next(t for t in tools if t.name == "run_nlp_column_task")
        result = json.loads(
            nlp_tool.invoke(
                {
                    "doc_id": "reviews_doc",
                    "column": "reviews",
                    "task": "sentiment",
                    "output_mode": "append_columns",
                }
            )
        )

        assert result["processed_rows"] == 7
        assert result["unique_text_inputs"] == 6
        assert result["source_row_count"] == 7
        assert result["accounted_rows"] == 7
        assert result["row_accounting_valid"] is True
        assert result["result_counts"] == {"negative": 2, "neutral": 1, "positive": 4}
        assert result["written_file"] == "reviews__analyst_sentiment.csv"
        assert result["artifact_row_count_valid"] is True
        assert result["derived_columns"] == ["sentiment_label", "sentiment_score"]
        assert result["preview_columns"] == ["row_index", "reviews", "sentiment_label", "sentiment_score"]
        assert result["preview_rows"][0]["sentiment_label"] == "positive"
        assert len(chat.calls) == 2
        written = session.workspace.root / "reviews__analyst_sentiment.csv"
        payload = written.read_text()
        assert "sentiment_label" in payload
        assert "sentiment_score" in payload
        assert "positive" in payload

    def test_missing_nlp_batch_outputs_are_counted_as_unknown(self, tmp_path):
        csv_file = tmp_path / "reviews.csv"
        _write_csv(csv_file, "reviews\nGreat\nBad\nOkay\n")

        session = self._workspace_session(tmp_path)
        stores = _make_stores(source_path=str(csv_file))
        settings = _make_settings()
        chat = _RecordingChatModel(["not json", "still not json"])

        from agentic_chatbot_next.providers import ProviderBundle
        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools

        tools = make_data_analyst_tools(
            stores,
            session,
            settings=settings,
            providers=ProviderBundle(chat=chat, judge=chat, embeddings=object()),
        )
        nlp_tool = next(t for t in tools if t.name == "run_nlp_column_task")
        result = json.loads(
            nlp_tool.invoke(
                {
                    "doc_id": "reviews_doc",
                    "column": "reviews",
                    "task": "sentiment",
                    "output_mode": "summary_only",
                    "batch_size": 3,
                }
            )
        )

        assert result["processed_rows"] == 3
        assert result["accounted_rows"] == 3
        assert result["row_accounting_valid"] is True
        assert result["unknown_unique_text_inputs"] == 3
        assert result["result_counts"] == {"unknown": 3}
        assert result["failed_batches"]

    @pytest.mark.parametrize("task", ["summarize", "keywords"])
    def test_missing_non_classification_nlp_outputs_are_reported_incomplete(self, tmp_path, task):
        csv_file = tmp_path / "reviews.csv"
        _write_csv(csv_file, "reviews\nGreat experience\nSlow support\n")

        session = self._workspace_session(tmp_path)
        stores = _make_stores(source_path=str(csv_file))
        settings = _make_settings()
        chat = _RecordingChatModel(["not json", "still not json"])

        from agentic_chatbot_next.providers import ProviderBundle
        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools

        tools = make_data_analyst_tools(
            stores,
            session,
            settings=settings,
            providers=ProviderBundle(chat=chat, judge=chat, embeddings=object()),
        )
        nlp_tool = next(t for t in tools if t.name == "run_nlp_column_task")
        result = json.loads(
            nlp_tool.invoke(
                {
                    "doc_id": "reviews_doc",
                    "column": "reviews",
                    "task": task,
                    "output_mode": "summary_only",
                    "batch_size": 2,
                }
            )
        )

        assert result["processed_rows"] == 2
        assert result["accounted_rows"] == 2
        assert result["row_accounting_valid"] is True
        assert result["unknown_unique_text_inputs"] == 0
        assert result["incomplete_unique_text_inputs"] == 2
        assert result["fallback_unique_text_inputs"] == 2
        assert result["partial_failure"] is True
        assert result["failed_batches"]
        assert any("incomplete" in warning for warning in result["warnings"])

    def test_repairs_malformed_json_from_llm(self, tmp_path):
        csv_file = tmp_path / "reviews.csv"
        _write_csv(csv_file, "reviews\nGreat experience\n")

        session = self._workspace_session(tmp_path)
        stores = _make_stores(source_path=str(csv_file))
        settings = _make_settings()
        chat = _RecordingChatModel(
            [
                "not valid json",
                json.dumps({"results": [{"item_id": "row_1", "label": "positive", "score": 0.98}]}),
            ]
        )

        from agentic_chatbot_next.providers import ProviderBundle
        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools

        tools = make_data_analyst_tools(
            stores,
            session,
            settings=settings,
            providers=ProviderBundle(chat=chat, judge=chat, embeddings=object()),
        )
        nlp_tool = next(t for t in tools if t.name == "run_nlp_column_task")
        result = json.loads(
            nlp_tool.invoke(
                {
                    "doc_id": "reviews_doc",
                    "column": "reviews",
                    "task": "sentiment",
                    "output_mode": "summary_only",
                }
            )
        )

        assert result["processed_rows"] == 1
        assert result["result_counts"]["positive"] == 1
        assert "summary_text" in result
        assert "Processed 1 rows" in result["summary_text"]
        assert len(chat.calls) == 2

    def test_accepts_dataset_and_sentiment_analysis_aliases(self, tmp_path):
        session = self._workspace_session(tmp_path)
        session.workspace.write_text("reviews.csv", "reviews\nGreat experience\n")
        stores = _make_stores()
        settings = _make_settings()
        chat = _RecordingChatModel(
            [
                json.dumps({"results": [{"item_id": "row_1", "label": "positive", "score": 0.98}]}),
            ]
        )

        from agentic_chatbot_next.providers import ProviderBundle
        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools

        tools = make_data_analyst_tools(
            stores,
            session,
            settings=settings,
            providers=ProviderBundle(chat=chat, judge=chat, embeddings=object()),
        )
        nlp_tool = next(t for t in tools if t.name == "run_nlp_column_task")
        result = json.loads(
            nlp_tool.invoke(
                {
                    "dataset": "reviews.csv",
                    "column": "reviews",
                    "task": "sentiment_analysis",
                    "output_mode": "summary_only",
                }
            )
        )

        assert result["task"] == "sentiment"
        assert result["doc_id"] == "reviews.csv"
        assert result["result_counts"]["positive"] == 1
        assert result["derived_columns"] == ["sentiment_label", "sentiment_score"]
        assert "summary_text" in result
        assert "positive: 1" in result["summary_text"]

    def test_summary_task_can_append_row_summary_column(self, tmp_path):
        csv_file = tmp_path / "reviews.csv"
        _write_csv(csv_file, "reviews\nGreat experience\nSlow support\n")

        session = self._workspace_session(tmp_path)
        stores = _make_stores(source_path=str(csv_file))
        settings = _make_settings()
        chat = _RecordingChatModel(
            [
                json.dumps(
                    {
                        "results": [
                            {"item_id": "row_1", "summary": "Customer had a great experience."},
                            {"item_id": "row_2", "summary": "Customer reported slow support."},
                        ]
                    }
                ),
            ]
        )

        from agentic_chatbot_next.providers import ProviderBundle
        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools

        tools = make_data_analyst_tools(
            stores,
            session,
            settings=settings,
            providers=ProviderBundle(chat=chat, judge=chat, embeddings=object()),
        )
        nlp_tool = next(t for t in tools if t.name == "run_nlp_column_task")
        result = json.loads(
            nlp_tool.invoke(
                {
                    "doc_id": "reviews_doc",
                    "column": "reviews",
                    "task": "summarize",
                    "output_mode": "append_columns",
                }
            )
        )

        assert result["processed_rows"] == 2
        assert result["written_file"] == "reviews__analyst_summarize.csv"
        assert result["derived_columns"] == ["row_summary"]
        assert result["preview_rows"][0]["row_summary"] == "Customer had a great experience."
        written = session.workspace.root / "reviews__analyst_summarize.csv"
        payload = written.read_text()
        assert "row_summary" in payload
        assert "Customer reported slow support." in payload

    def test_return_file_registers_download_artifact(self, tmp_path):
        session = self._workspace_session(tmp_path)
        session.workspace.write_text("analysis.txt", "done")
        session.scratchpad["last_output_file"] = "analysis.txt"
        stores = _make_stores()
        settings = _make_settings()

        from agentic_chatbot_next.tools.data_analyst_tools import make_data_analyst_tools

        tools = make_data_analyst_tools(stores, session, settings=settings)
        return_tool = next(t for t in tools if t.name == "return_file")
        result = json.loads(return_tool.invoke({}))

        assert result["filename"] == "analysis.txt"
        assert result["download_url"].startswith("/v1/files/")
        assert result["download_id"] in session.metadata["downloads"]
