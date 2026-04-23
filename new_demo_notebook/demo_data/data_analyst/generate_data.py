from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parent

REGIONS = ["North", "South", "East", "West", "Central", "Enterprise"]
PRODUCT_LINES = ["Core", "Analytics", "Automation", "Billing"]
CHANNELS = ["Email", "Web", "App", "Partner", "Retail"]
SEGMENTS = ["SMB", "Mid-Market", "Enterprise"]
MANAGERS = ["A. Rivera", "J. Patel", "N. Brooks"]

REVIEW_TEMPLATES = [
    ("Lightning-fast checkout and zero issues so far.", 5, "positive", "clear praise"),
    ("Support never replied and I lost a day of work.", 1, "negative", "clear frustration"),
    ("It works, but only after the third refresh.", 3, "neutral", "mixed experience"),
    ("Great, another update that moved every button I use.", 2, "negative", "sarcasm"),
    ("Honestly pretty solid for a first release.", 4, "positive", "mild praise"),
    ("I cannot tell if this is brilliant or broken.", 3, "neutral", "ambiguous"),
    ("Saved my team hours this week.", 5, "positive", "workflow win"),
    ("Billing is a mess and the invoice export failed twice.", 1, "negative", "billing pain"),
    ("It is fine, I guess. Nothing special.", 3, "neutral", "flat tone"),
    ("The new dashboard looks nice, but reports are slower now.", 3, "neutral", "mixed tradeoff"),
    ("Best onboarding experience we have had in years.", 5, "positive", "strong praise"),
    ("The product is okay, the documentation is not.", 2, "negative", "docs complaint"),
    ("Not bad once you learn where everything lives.", 4, "positive", "qualified praise"),
    ("I expected worse, so this is a pleasant surprise.", 4, "positive", "surprised praise"),
    ("", "", "", "blank review"),
    ("The rollout was smooth and the data matched our old system.", 5, "positive", "migration success"),
    ("Maybe it is user error, but I still cannot find the export button.", 2, "negative", "confused frustration"),
    ("Nothing crashed, nothing wowed me.", 3, "neutral", "balanced"),
    ("This solved one problem and created two new ones.", 2, "negative", "tradeoff"),
    ("Fast, clear, and exactly what we needed.", 5, "positive", "strong praise"),
]

TICKET_MESSAGES = [
    ("The login page keeps spinning and we are locked out.", "Access", "negative"),
    ("Need help understanding why the weekly report is blank.", "Reporting", "neutral"),
    ("Everything is working now, thanks for the fast turnaround.", "Support", "positive"),
    ("This is the third outage alert today and nobody has called us back.", "Reliability", "negative"),
    ("Could someone confirm whether the sync finished correctly?", "Integrations", "neutral"),
    ("Our finance team loves the new export layout.", "Billing", "positive"),
    ("The workflow is slower after the patch and agents are timing out.", "Automation", "negative"),
    ("I am not sure whether the SLA breach warning is accurate.", "Operations", "neutral"),
]


def _write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _write_sheet(workbook: Workbook, title: str, rows: list[dict[str, object]]) -> None:
    worksheet = workbook.active if workbook.active.title == "Sheet" and workbook.active.max_row == 1 and workbook.active.max_column == 1 else workbook.create_sheet()
    worksheet.title = title
    if not rows:
        return
    headers = list(rows[0].keys())
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])


def _build_customer_reviews_rows() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for index in range(100):
        template = REVIEW_TEMPLATES[index % len(REVIEW_TEMPLATES)]
        region = REGIONS[index % len(REGIONS)]
        product_line = PRODUCT_LINES[index % len(PRODUCT_LINES)]
        channel = CHANNELS[index % len(CHANNELS)]
        segment = SEGMENTS[index % len(SEGMENTS)]
        rows.append(
            {
                "review_id": f"REV-{index + 1:03d}",
                "customer_segment": segment,
                "region": region,
                "product_line": product_line,
                "channel": channel,
                "rating": template[1],
                "reviews": template[0],
                "expected_sentiment_hint": template[2],
                "annotation_note": template[3],
            }
        )
    return rows


def _build_customer_reviews_multisheet(path: Path, review_rows: list[dict[str, object]]) -> None:
    workbook = Workbook()
    raw_reviews = [
        {
            "review_id": row["review_id"],
            "region": row["region"],
            "channel": row["channel"],
            "rating": row["rating"],
            "reviews": row["reviews"],
        }
        for row in review_rows[:30]
    ]
    metadata = [
        {"key": "dataset", "value": "customer_reviews_multisheet"},
        {"key": "primary_sheet", "value": "raw_reviews"},
        {"key": "purpose", "value": "row-level sentiment labeling and workbook mutation tests"},
        {"key": "blank_review_behavior", "value": "blank cells should remain blank after classification"},
    ]
    expected_samples = [
        {
            "review_id": row["review_id"],
            "reviews": row["reviews"],
            "expected_label": row["expected_sentiment_hint"],
            "notes": row["annotation_note"],
        }
        for row in review_rows[:12]
    ]
    _write_sheet(workbook, "raw_reviews", raw_reviews)
    _write_sheet(workbook, "metadata", metadata)
    _write_sheet(workbook, "expected_sentiment_samples", expected_samples)
    workbook.save(path)


def _build_sales_performance(path: Path) -> None:
    workbook = Workbook()
    sales_rows: list[dict[str, object]] = []
    reps = [
        ("Morgan Lee", "North", "A. Rivera"),
        ("Priya Shah", "South", "A. Rivera"),
        ("Derek Hall", "East", "J. Patel"),
        ("Sofia Kim", "West", "J. Patel"),
        ("Elena Cruz", "Central", "N. Brooks"),
        ("Noah James", "Enterprise", "N. Brooks"),
    ]
    for month_number in range(1, 7):
        for rep_index, (rep_name, region, manager) in enumerate(reps):
            base_spend = 18000 + month_number * 1400 + rep_index * 2300
            pipeline = base_spend * 3.2 + rep_index * 12000
            revenue = base_spend * 2.1 + month_number * 7500 + rep_index * 4500
            sales_rows.append(
                {
                    "month": f"2026-{month_number:02d}-01",
                    "region": region,
                    "rep_name": rep_name,
                    "manager": manager,
                    "segment": SEGMENTS[rep_index % len(SEGMENTS)],
                    "deals_closed": 8 + month_number + rep_index,
                    "units_sold": 55 + month_number * 4 + rep_index * 3,
                    "avg_discount_pct": round(4.5 + rep_index * 0.8 + month_number * 0.2, 2),
                    "marketing_spend_usd": round(base_spend, 2),
                    "pipeline_value_usd": round(pipeline, 2),
                    "revenue_usd": round(revenue, 2),
                    "renewal_rate_pct": round(84 + rep_index * 1.5 - month_number * 0.4, 2),
                }
            )
    target_rows = [
        {
            "region": region,
            "monthly_target_revenue_usd": 62000 + idx * 9000,
            "monthly_target_pipeline_usd": 170000 + idx * 20000,
            "quota_attainment_target_pct": 96 + idx,
        }
        for idx, region in enumerate(REGIONS)
    ]
    rep_rows = [
        {
            "rep_name": rep_name,
            "region": region,
            "manager": manager,
            "tenure_months": 10 + idx * 7,
            "focus_segment": SEGMENTS[idx % len(SEGMENTS)],
        }
        for idx, (rep_name, region, manager) in enumerate(reps)
    ]
    _write_sheet(workbook, "sales_data", sales_rows)
    _write_sheet(workbook, "targets", target_rows)
    _write_sheet(workbook, "rep_directory", rep_rows)
    workbook.save(path)


def _build_marketing_leads() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    campaigns = ["Spring Push", "Retention Reactivation", "Webinar", "Partner Boost"]
    owners = ["J. Patel", "R. Flores", "K. Young", "A. White"]
    for index in range(48):
        spend = "" if index % 11 == 0 else round(1200 + index * 87.5, 2)
        conversions = "" if index % 9 == 0 else 4 + (index % 6)
        clicks = 90 + index * 7
        rows.append(
            {
                "lead_id": f"LEAD-{index + 1:03d}",
                "created_at": f"2026-03-{(index % 28) + 1:02d}",
                "region": REGIONS[index % len(REGIONS)],
                "channel": CHANNELS[index % len(CHANNELS)],
                "campaign": campaigns[index % len(campaigns)],
                "spend_usd": spend,
                "impressions": 5000 + index * 260,
                "clicks": clicks,
                "leads": 12 + (index % 10),
                "conversions": conversions,
                "revenue_usd": round(4200 + index * 310.5, 2),
                "landing_page_score": round(68 + (index % 8) * 2.5, 1),
                "owner": owners[index % len(owners)],
                "notes": "Missing spend" if spend == "" else ("Delayed CRM sync" if index % 7 == 0 else ""),
            }
        )
    return rows


def _build_support_tickets(path: Path) -> None:
    workbook = Workbook()
    ticket_rows: list[dict[str, object]] = []
    team_rows = [
        {"team": "Platform", "manager": "A. Rivera", "region": "North"},
        {"team": "Support", "manager": "J. Patel", "region": "South"},
        {"team": "Reliability", "manager": "N. Brooks", "region": "West"},
    ]
    qa_rows: list[dict[str, object]] = []
    priorities = ["P1", "P2", "P3", "P4"]
    teams = ["Platform", "Support", "Reliability"]
    for index in range(40):
        message, product_area, expected_sentiment = TICKET_MESSAGES[index % len(TICKET_MESSAGES)]
        resolution_hours = round(2.5 + (index % 9) * 1.75, 2)
        csat = "" if index % 6 == 0 else round(2.5 + (index % 5) * 0.6, 1)
        ticket_rows.append(
            {
                "ticket_id": f"TCK-{index + 1:03d}",
                "opened_at": f"2026-02-{(index % 27) + 1:02d} 09:{(index * 7) % 60:02d}",
                "priority": priorities[index % len(priorities)],
                "team": teams[index % len(teams)],
                "product_area": product_area,
                "sla_hours": 4 if index % 4 == 0 else 8 if index % 3 == 0 else 24,
                "first_response_minutes": 8 + (index % 12) * 6,
                "resolution_hours": resolution_hours,
                "csat_score": csat,
                "customer_message": message,
                "agent_summary": "Escalated to engineering" if index % 5 == 0 else "Resolved in support",
            }
        )
        if index < 12:
            qa_rows.append(
                {
                    "ticket_id": f"TCK-{index + 1:03d}",
                    "customer_message": message,
                    "expected_sentiment": expected_sentiment,
                    "expected_urgency": "high" if priorities[index % len(priorities)] in {"P1", "P2"} else "normal",
                }
            )
    _write_sheet(workbook, "tickets", ticket_rows)
    _write_sheet(workbook, "team_roster", team_rows)
    _write_sheet(workbook, "qa_labels", qa_rows)
    workbook.save(path)


def main() -> None:
    ROOT.mkdir(parents=True, exist_ok=True)
    review_rows = _build_customer_reviews_rows()
    _write_csv(
        ROOT / "customer_reviews_100.csv",
        review_rows,
        [
            "review_id",
            "customer_segment",
            "region",
            "product_line",
            "channel",
            "rating",
            "reviews",
            "expected_sentiment_hint",
            "annotation_note",
        ],
    )
    _build_customer_reviews_multisheet(ROOT / "customer_reviews_multisheet.xlsx", review_rows)
    _build_sales_performance(ROOT / "sales_performance.xlsx")
    _write_csv(
        ROOT / "marketing_leads.csv",
        _build_marketing_leads(),
        [
            "lead_id",
            "created_at",
            "region",
            "channel",
            "campaign",
            "spend_usd",
            "impressions",
            "clicks",
            "leads",
            "conversions",
            "revenue_usd",
            "landing_page_score",
            "owner",
            "notes",
        ],
    )
    _build_support_tickets(ROOT / "support_tickets.xlsx")


if __name__ == "__main__":
    main()
