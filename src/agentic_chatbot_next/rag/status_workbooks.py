from __future__ import annotations

import datetime as dt
import math
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Iterable, Sequence


STATUS_WORKBOOK_METADATA_VERSION = "status_workbook_v1"

STATUS_DOMAINS: tuple[str, ...] = (
    "risk",
    "issue",
    "action",
    "schedule",
    "budget",
    "cdrl",
    "requirement",
    "test_event",
    "milestone",
    "status",
)

_DOMAIN_TERMS: dict[str, tuple[str, ...]] = {
    "risk": ("risk", "risks", "mitigation", "probability", "likelihood", "impact", "exposure"),
    "issue": ("issue", "issues", "blocker", "problem", "impediment", "defect", "constraint"),
    "action": ("action", "actions", "action item", "task", "todo", "follow up", "ai"),
    "schedule": (
        "schedule",
        "planned",
        "forecast",
        "baseline",
        "start",
        "finish",
        "complete",
        "completion",
        "date",
        "slip",
        "variance",
    ),
    "budget": (
        "budget",
        "cost",
        "amount",
        "funding",
        "spend",
        "actual",
        "forecast",
        "eac",
        "etc",
        "variance",
        "overrun",
    ),
    "cdrl": ("cdrl", "data item", "deliverable", "did", "submission", "delivery", "deliverables"),
    "requirement": ("requirement", "requirements", "req", "shall", "verification method", "compliance"),
    "test_event": (
        "test",
        "event",
        "test event",
        "verification",
        "validation",
        "qualification",
        "trr",
        "readiness",
    ),
    "milestone": (
        "milestone",
        "milestones",
        "review",
        "srr",
        "pdr",
        "cdr",
        "trr",
        "orr",
        "mrr",
        "ffr",
        "delivery",
    ),
    "status": (
        "status",
        "state",
        "health",
        "rag",
        "condition",
        "open",
        "closed",
        "current",
        "latest",
        "approved",
        "complete",
        "done",
    ),
}

_COLUMN_ROLE_TERMS: dict[str, tuple[str, ...]] = {
    "identifier": ("id", "identifier", "number", "no", "key", "req id", "risk id", "issue id", "action id"),
    "title": ("title", "name", "subject", "item", "milestone", "event", "deliverable", "requirement"),
    "description": ("description", "desc", "summary", "details", "narrative", "notes", "comment", "comments"),
    "status": ("status", "state", "health", "rag", "condition", "approval", "approved"),
    "owner": ("owner", "responsible", "assignee", "assigned to", "poc", "lead", "manager", "prime"),
    "date": ("date", "created", "updated", "approved", "as of"),
    "start_date": ("start", "start date", "begin", "planned start", "baseline start"),
    "due_date": (
        "due",
        "due date",
        "target",
        "need date",
        "deadline",
        "delivery",
        "forecast finish",
        "planned finish",
        "completion",
        "complete by",
    ),
    "finish_date": ("finish", "finish date", "end", "complete", "completion", "closed"),
    "severity": ("severity", "criticality", "impact"),
    "priority": ("priority", "rank"),
    "probability": ("probability", "likelihood"),
    "impact": ("impact", "consequence"),
    "mitigation": ("mitigation", "mitigate", "response", "handling", "contingency"),
    "budget_amount": ("budget", "cost", "amount", "funding", "spend", "actual", "forecast", "eac", "etc"),
    "variance": ("variance", "delta", "slip", "overrun", "underrun"),
}

_ROLE_DOMAIN_HINTS: dict[str, str] = {
    "mitigation": "risk",
    "probability": "risk",
    "impact": "risk",
    "budget_amount": "budget",
    "variance": "schedule",
    "due_date": "schedule",
    "start_date": "schedule",
    "finish_date": "schedule",
}

_MILESTONE_ACRONYM_RE = re.compile(r"\b(?:SRR|PDR|CDR|TRR|ORR|MRR|FFR)\b", re.IGNORECASE)
_MONEY_RE = re.compile(r"(?:\$|\bUSD\b|\b\d+(?:\.\d+)?\s*(?:k|m|million|thousand)\b)", re.IGNORECASE)


@dataclass(frozen=True)
class ColumnRole:
    column_index: int
    column_letter: str
    header: str
    role: str
    domain: str = ""
    confidence: float = 0.0
    matched_terms: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class StatusSourceRef:
    doc_id: str = ""
    title: str = ""
    source_path: str = ""
    sheet_name: str = ""
    row_start: int | None = None
    row_end: int | None = None
    cell_range: str = ""
    columns: list[str] = field(default_factory=list)
    confidence: float = 0.0

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class StatusRecord:
    domain: str
    summary: str
    record_id: str = ""
    status: str = ""
    owner: str = ""
    date: str = ""
    due_date: str = ""
    amount: str = ""
    severity: str = ""
    priority: str = ""
    values: dict[str, Any] = field(default_factory=dict)
    source_ref: StatusSourceRef = field(default_factory=StatusSourceRef)
    confidence: float = 0.0

    def to_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        payload["source_ref"] = self.source_ref.to_dict()
        return payload


@dataclass(frozen=True)
class SheetProfile:
    sheet_name: str
    visible: bool = True
    header_row: int | None = None
    row_count: int = 0
    column_count: int = 0
    cell_range: str = ""
    headers: list[str] = field(default_factory=list)
    named_tables: list[dict[str, Any]] = field(default_factory=list)
    merged_cells: list[str] = field(default_factory=list)
    column_roles: list[ColumnRole] = field(default_factory=list)
    domain_tags: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        payload["column_roles"] = [item.to_dict() for item in self.column_roles]
        return payload


@dataclass(frozen=True)
class WorkbookProfile:
    workbook_name: str
    source_path: str = ""
    file_type: str = ""
    sheets: list[SheetProfile] = field(default_factory=list)
    domain_counts: dict[str, int] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    metadata_version: str = STATUS_WORKBOOK_METADATA_VERSION

    def to_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        payload["sheets"] = [item.to_dict() for item in self.sheets]
        return payload


@dataclass(frozen=True)
class _SheetRows:
    sheet_name: str
    visible: bool
    rows: list[tuple[int, tuple[Any, ...]]]
    row_count: int = 0
    column_count: int = 0
    cell_range: str = ""
    named_tables: list[dict[str, Any]] = field(default_factory=list)
    merged_cells: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def profile_workbook(path: Path | str) -> WorkbookProfile:
    """Return deterministic workbook structure and status-column metadata."""

    resolved = Path(path)
    warnings: list[str] = []
    sheets: list[SheetProfile] = []
    domain_counts: dict[str, int] = {}
    try:
        sheet_rows = _load_sheet_rows(resolved)
    except Exception as exc:
        return WorkbookProfile(
            workbook_name=resolved.name,
            source_path=str(resolved),
            file_type=resolved.suffix.lstrip(".").lower(),
            warnings=[f"Workbook profile failed: {exc}"],
        )

    for sheet in sheet_rows:
        profile = _profile_sheet(sheet)
        sheets.append(profile)
        for domain in profile.domain_tags:
            domain_counts[domain] = domain_counts.get(domain, 0) + 1
        warnings.extend(profile.warnings)
    return WorkbookProfile(
        workbook_name=resolved.name,
        source_path=str(resolved),
        file_type=resolved.suffix.lstrip(".").lower(),
        sheets=sheets,
        domain_counts=dict(sorted(domain_counts.items())),
        warnings=_dedupe(warnings),
    )


def classify_headers(headers: Sequence[Any]) -> list[ColumnRole]:
    roles: list[ColumnRole] = []
    for index, raw_header in enumerate(headers):
        header = _cell_text(raw_header) or f"Column {_column_label(index)}"
        normalized = _normalize_text(header)
        for role, terms in _COLUMN_ROLE_TERMS.items():
            matches = [term for term in terms if _term_matches(normalized, term)]
            if not matches:
                continue
            domain = _ROLE_DOMAIN_HINTS.get(role, "")
            domain_matches = _domains_from_text(header)
            if not domain and domain_matches:
                domain = domain_matches[0]
            confidence = min(0.95, 0.55 + (len(matches) * 0.08) + (0.08 if domain else 0.0))
            roles.append(
                ColumnRole(
                    column_index=index,
                    column_letter=_column_label(index),
                    header=header,
                    role=role,
                    domain=domain,
                    confidence=round(confidence, 2),
                    matched_terms=matches[:6],
                )
            )
    return roles


def status_metadata_for_sheet(headers: Sequence[Any]) -> dict[str, Any]:
    normalized_headers = _normalize_headers(headers)
    column_roles = classify_headers(normalized_headers)
    domains = _domains_from_headers(normalized_headers, column_roles)
    return {
        "metadata_version": STATUS_WORKBOOK_METADATA_VERSION,
        "headers": list(normalized_headers),
        "column_roles": [role.to_dict() for role in column_roles],
        "status_domains": domains,
    }


def status_metadata_for_row(
    headers: Sequence[Any],
    values: Sequence[Any],
    *,
    row_number: int,
) -> dict[str, Any]:
    normalized_headers = _normalize_headers(headers)
    text_values = [_cell_text(value) for value in values]
    column_roles = classify_headers(normalized_headers)
    domains = _domains_for_row(normalized_headers, text_values, column_roles)
    if not domains:
        return {}
    key_values = _record_values(normalized_headers, text_values, column_roles)
    return {
        "metadata_version": STATUS_WORKBOOK_METADATA_VERSION,
        "row_number": int(row_number),
        "status_domains": domains,
        "key_values": key_values,
        "column_roles": [role.to_dict() for role in column_roles],
        "record_summary": _row_summary(domains[0], key_values),
    }


def detect_status_domains_for_query(query: str) -> list[str]:
    text = str(query or "")
    normalized = _normalize_text(text)
    domains: list[str] = []
    for domain in STATUS_DOMAINS:
        if any(_term_matches(normalized, term) for term in _DOMAIN_TERMS.get(domain, ())):
            domains.append(domain)
    if _MILESTONE_ACRONYM_RE.search(text):
        _add_unique(domains, "milestone")
    if re.search(r"\b(?:when|date|deadline|due|late|slip|current|approved|latest)\b", text, re.I):
        _add_unique(domains, "schedule")
    if re.search(r"\b(?:open|closed|blocked|green|yellow|red|amber|complete|in progress)\b", text, re.I):
        _add_unique(domains, "status")
    if _MONEY_RE.search(text):
        _add_unique(domains, "budget")
    return [domain for domain in STATUS_DOMAINS if domain in set(domains)]


def query_has_status_intent(query: str) -> bool:
    if detect_status_domains_for_query(query):
        return True
    return bool(
        re.search(
            r"\b(status|risk|issue|action|milestone|budget|cdrl|requirement|test event|schedule|owner|due date)\b",
            str(query or ""),
            re.I,
        )
    )


def extract_status_records(
    path: Path | str,
    *,
    domains: Sequence[str] | None = None,
    sheet_name: str = "",
    status_filter: str = "",
    doc_id: str = "",
    title: str = "",
    source_path: str = "",
    max_records: int = 100,
) -> list[StatusRecord]:
    resolved = Path(path)
    requested_domains = _normalize_domains(domains)
    clean_sheet = str(sheet_name or "").strip()
    clean_status = _normalize_text(status_filter)
    records: list[StatusRecord] = []
    sheet_rows = _load_sheet_rows(resolved)
    for sheet in sheet_rows:
        if clean_sheet and sheet.sheet_name != clean_sheet:
            continue
        if not sheet.visible and not clean_sheet:
            continue
        trimmed = _trim_rows(sheet.rows)
        if not trimmed:
            continue
        header_pos = _detect_header_position(trimmed)
        header_row, raw_headers = trimmed[header_pos]
        headers = _normalize_headers(raw_headers)
        column_roles = classify_headers(headers)
        for row_number, raw_values in trimmed[header_pos + 1 :]:
            text_values = [_cell_text(value) for value in raw_values]
            if not any(text_values):
                continue
            row_domains = _domains_for_row(headers, text_values, column_roles)
            if requested_domains:
                row_domains = [domain for domain in requested_domains if domain in set(row_domains)]
            if not row_domains:
                continue
            values = _record_values(headers, text_values, column_roles)
            if clean_status:
                row_status = _normalize_text(values.get("status") or "")
                if clean_status not in row_status:
                    continue
            domain = row_domains[0]
            max_col = _column_label(max(0, len(headers) - 1))
            source_ref = StatusSourceRef(
                doc_id=str(doc_id or ""),
                title=str(title or resolved.name),
                source_path=str(source_path or resolved),
                sheet_name=sheet.sheet_name,
                row_start=int(row_number),
                row_end=int(row_number),
                cell_range=f"A{row_number}:{max_col}{row_number}",
                columns=_source_columns(headers, text_values, column_roles),
                confidence=_row_confidence(values, row_domains),
            )
            records.append(
                StatusRecord(
                    domain=domain,
                    summary=_row_summary(domain, values),
                    record_id=str(values.get("identifier") or ""),
                    status=str(values.get("status") or ""),
                    owner=str(values.get("owner") or ""),
                    date=str(values.get("date") or ""),
                    due_date=str(values.get("due_date") or values.get("finish_date") or ""),
                    amount=str(values.get("budget_amount") or ""),
                    severity=str(values.get("severity") or ""),
                    priority=str(values.get("priority") or ""),
                    values=values,
                    source_ref=source_ref,
                    confidence=source_ref.confidence,
                )
            )
            if len(records) >= max(1, int(max_records or 100)):
                return records
    return records


def extract_risks(path: Path | str, **kwargs: Any) -> list[StatusRecord]:
    return extract_status_records(path, domains=["risk"], **kwargs)


def extract_issues(path: Path | str, **kwargs: Any) -> list[StatusRecord]:
    return extract_status_records(path, domains=["issue"], **kwargs)


def extract_actions(path: Path | str, **kwargs: Any) -> list[StatusRecord]:
    return extract_status_records(path, domains=["action"], **kwargs)


def extract_schedules(path: Path | str, **kwargs: Any) -> list[StatusRecord]:
    return extract_status_records(path, domains=["schedule"], **kwargs)


def extract_budgets(path: Path | str, **kwargs: Any) -> list[StatusRecord]:
    return extract_status_records(path, domains=["budget"], **kwargs)


def extract_cdrls(path: Path | str, **kwargs: Any) -> list[StatusRecord]:
    return extract_status_records(path, domains=["cdrl"], **kwargs)


def extract_requirements(path: Path | str, **kwargs: Any) -> list[StatusRecord]:
    return extract_status_records(path, domains=["requirement"], **kwargs)


def extract_test_events(path: Path | str, **kwargs: Any) -> list[StatusRecord]:
    return extract_status_records(path, domains=["test_event"], **kwargs)


def extract_milestones(path: Path | str, **kwargs: Any) -> list[StatusRecord]:
    return extract_status_records(path, domains=["milestone"], **kwargs)


def _load_sheet_rows(path: Path) -> list[_SheetRows]:
    suffix = path.suffix.lower()
    if suffix == ".xls":
        return _load_xls_rows(path)
    return _load_xlsx_rows(path)


def _load_xlsx_rows(path: Path) -> list[_SheetRows]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=False, data_only=True)
    sheets: list[_SheetRows] = []
    try:
        for sheet in workbook.worksheets:
            rows: list[tuple[int, tuple[Any, ...]]] = []
            for row_index, cells in enumerate(sheet.iter_rows(values_only=True), start=1):
                if bool(getattr(sheet.row_dimensions[row_index], "hidden", False)):
                    continue
                rows.append((row_index, tuple(_trim_trailing_empty(cells))))
            named_tables: list[dict[str, Any]] = []
            for table in list(getattr(sheet, "tables", {}).values()):
                named_tables.append(
                    {
                        "name": str(getattr(table, "displayName", "") or getattr(table, "name", "") or ""),
                        "ref": str(getattr(table, "ref", "") or ""),
                    }
                )
            sheets.append(
                _SheetRows(
                    sheet_name=sheet.title,
                    visible=str(getattr(sheet, "sheet_state", "visible") or "visible") == "visible",
                    rows=rows,
                    row_count=int(sheet.max_row or 0),
                    column_count=int(sheet.max_column or 0),
                    cell_range=str(sheet.calculate_dimension() or ""),
                    named_tables=[item for item in named_tables if item.get("name") or item.get("ref")],
                    merged_cells=[str(item) for item in sheet.merged_cells.ranges],
                )
            )
    finally:
        workbook.close()
    return sheets


def _load_xls_rows(path: Path) -> list[_SheetRows]:
    import pandas as pd

    workbook = pd.ExcelFile(path)
    sheets: list[_SheetRows] = []
    for name in workbook.sheet_names:
        frame = workbook.parse(sheet_name=name, header=None)
        rows = [
            (index + 1, tuple(values))
            for index, values in enumerate(frame.itertuples(index=False, name=None))
        ]
        row_count, column_count = frame.shape
        sheets.append(
            _SheetRows(
                sheet_name=str(name),
                visible=True,
                rows=rows,
                row_count=int(row_count or 0),
                column_count=int(column_count or 0),
                cell_range=f"A1:{_column_label(max(0, int(column_count or 1) - 1))}{int(row_count or 1)}",
            )
        )
    return sheets


def _profile_sheet(sheet: _SheetRows) -> SheetProfile:
    trimmed = _trim_rows(sheet.rows)
    warnings = list(sheet.warnings)
    if not trimmed:
        return SheetProfile(
            sheet_name=sheet.sheet_name,
            visible=sheet.visible,
            row_count=sheet.row_count,
            column_count=sheet.column_count,
            cell_range=sheet.cell_range,
            named_tables=list(sheet.named_tables),
            merged_cells=list(sheet.merged_cells),
            warnings=warnings,
        )
    header_pos = _detect_header_position(trimmed)
    header_row, raw_headers = trimmed[header_pos]
    headers = _normalize_headers(raw_headers)
    column_roles = classify_headers(headers)
    domain_tags = _domains_from_headers(headers, column_roles)
    for _, raw_values in trimmed[header_pos + 1 :]:
        row_domains = _domains_for_row(headers, [_cell_text(value) for value in raw_values], column_roles)
        for domain in row_domains:
            _add_unique(domain_tags, domain)
    return SheetProfile(
        sheet_name=sheet.sheet_name,
        visible=sheet.visible,
        header_row=int(header_row),
        row_count=sheet.row_count,
        column_count=max(sheet.column_count, len(headers)),
        cell_range=sheet.cell_range,
        headers=headers,
        named_tables=list(sheet.named_tables),
        merged_cells=list(sheet.merged_cells),
        column_roles=column_roles,
        domain_tags=[domain for domain in STATUS_DOMAINS if domain in set(domain_tags)],
        warnings=warnings,
    )


def _trim_rows(rows: Sequence[tuple[int, tuple[Any, ...]]]) -> list[tuple[int, tuple[Any, ...]]]:
    return [(row_number, values) for row_number, values in rows if any(_cell_text(value) for value in values)]


def _trim_trailing_empty(values: Sequence[Any]) -> list[Any]:
    items = list(values)
    while items and not _cell_text(items[-1]):
        items.pop()
    return items


def _detect_header_position(rows: Sequence[tuple[int, tuple[Any, ...]]]) -> int:
    best_index = 0
    best_score = -1.0
    for index, (_row_number, values) in enumerate(rows[:12]):
        score = _header_score(values)
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def _header_score(values: Sequence[Any]) -> float:
    cells = [_cell_text(value) for value in values]
    nonempty = [cell for cell in cells if cell]
    if len(nonempty) < 2:
        return -1.0
    normalized = [_normalize_text(cell) for cell in nonempty]
    unique = len(set(normalized))
    alpha = sum(1 for cell in nonempty if any(ch.isalpha() for ch in cell))
    role_hits = sum(1 for role in classify_headers(nonempty) if role.confidence >= 0.55)
    domain_hits = len(_domains_from_text(" ".join(nonempty)))
    return (len(nonempty) * 1.2) + (unique * 0.6) + (alpha * 0.5) + (role_hits * 2.2) + (domain_hits * 1.5)


def _normalize_headers(values: Sequence[Any]) -> list[str]:
    headers: list[str] = []
    used: set[str] = set()
    for index, raw in enumerate(values):
        text = _cell_text(raw) or f"Column {_column_label(index)}"
        candidate = text
        suffix = 2
        while candidate.casefold() in used:
            candidate = f"{text} {suffix}"
            suffix += 1
        used.add(candidate.casefold())
        headers.append(candidate)
    return headers


def _domains_from_headers(headers: Sequence[str], column_roles: Sequence[ColumnRole]) -> list[str]:
    domains: list[str] = []
    header_text = " ".join(headers)
    for domain in _domains_from_text(header_text):
        _add_unique(domains, domain)
    for role in column_roles:
        if role.domain:
            _add_unique(domains, role.domain)
        if role.role == "status":
            _add_unique(domains, "status")
    return [domain for domain in STATUS_DOMAINS if domain in set(domains)]


def _domains_for_row(headers: Sequence[str], values: Sequence[str], column_roles: Sequence[ColumnRole]) -> list[str]:
    domains = _domains_from_headers(headers, column_roles)
    combined = " ".join([*headers, *[str(value or "") for value in values if str(value or "").strip()]])
    for domain in _domains_from_text(combined):
        _add_unique(domains, domain)
    if _MILESTONE_ACRONYM_RE.search(combined):
        _add_unique(domains, "milestone")
    if _MONEY_RE.search(combined):
        _add_unique(domains, "budget")
    return [domain for domain in STATUS_DOMAINS if domain in set(domains)]


def _domains_from_text(value: str) -> list[str]:
    normalized = _normalize_text(value)
    domains: list[str] = []
    for domain in STATUS_DOMAINS:
        if any(_term_matches(normalized, term) for term in _DOMAIN_TERMS.get(domain, ())):
            domains.append(domain)
    return domains


def _record_values(headers: Sequence[str], values: Sequence[str], column_roles: Sequence[ColumnRole]) -> dict[str, Any]:
    by_role: dict[str, Any] = {}
    roles_by_col: dict[int, list[ColumnRole]] = {}
    for role in column_roles:
        roles_by_col.setdefault(role.column_index, []).append(role)
    for index, header in enumerate(headers):
        value = values[index] if index < len(values) else ""
        if not value:
            continue
        for role in roles_by_col.get(index, []):
            by_role.setdefault(role.role, value)
        by_role.setdefault(header, value)
    compact: dict[str, Any] = {}
    for key in (
        "identifier",
        "title",
        "description",
        "status",
        "owner",
        "date",
        "start_date",
        "due_date",
        "finish_date",
        "severity",
        "priority",
        "probability",
        "impact",
        "mitigation",
        "budget_amount",
        "variance",
    ):
        if by_role.get(key):
            compact[key] = by_role[key]
    for header, value in zip(headers, values):
        if value and header not in compact:
            compact[header] = value
        if len(compact) >= 18:
            break
    return compact


def _source_columns(headers: Sequence[str], values: Sequence[str], column_roles: Sequence[ColumnRole]) -> list[str]:
    role_columns = {role.column_index for role in column_roles if role.role in {"identifier", "title", "status", "owner", "date", "due_date", "finish_date", "budget_amount", "variance", "severity", "priority"}}
    columns: list[str] = []
    for index, header in enumerate(headers):
        has_value = bool(values[index].strip()) if index < len(values) else False
        if index in role_columns or (has_value and len(columns) < 8):
            _add_unique(columns, header)
    return columns[:12]


def _row_summary(domain: str, values: dict[str, Any]) -> str:
    label = str(values.get("title") or values.get("description") or values.get("identifier") or domain).strip()
    parts = [f"{domain.replace('_', ' ').title()}: {label}"]
    for key, label_name in (
        ("status", "status"),
        ("owner", "owner"),
        ("due_date", "due"),
        ("finish_date", "finish"),
        ("date", "date"),
        ("budget_amount", "amount"),
        ("variance", "variance"),
        ("severity", "severity"),
        ("priority", "priority"),
    ):
        value = str(values.get(key) or "").strip()
        if value:
            parts.append(f"{label_name} {value}")
    return "; ".join(parts[:7])


def _row_confidence(values: dict[str, Any], domains: Sequence[str]) -> float:
    confidence = 0.55
    if domains:
        confidence += 0.1
    for key in ("status", "owner", "due_date", "date", "budget_amount", "severity", "priority"):
        if values.get(key):
            confidence += 0.04
    return round(min(confidence, 0.92), 2)


def _normalize_domains(domains: Sequence[str] | None) -> list[str]:
    normalized: list[str] = []
    for raw in domains or []:
        clean = str(raw or "").strip().lower().replace("-", "_").replace(" ", "_")
        if clean == "test":
            clean = "test_event"
        if clean in STATUS_DOMAINS:
            normalized.append(clean)
    return [domain for domain in STATUS_DOMAINS if domain in set(normalized)]


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        import pandas as pd

        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
        if value.is_integer():
            return str(int(value))
    return str(value).strip()


def _normalize_text(value: Any) -> str:
    return " ".join(re.findall(r"[a-z0-9]+", str(value or "").lower()))


def _term_matches(normalized_text: str, term: str) -> bool:
    normalized_term = _normalize_text(term)
    if not normalized_text or not normalized_term:
        return False
    text = f" {normalized_text} "
    term_text = f" {normalized_term} "
    return normalized_text == normalized_term or term_text in text


def _column_label(index: int) -> str:
    index = int(index)
    letters = ""
    while True:
        index, remainder = divmod(index, 26)
        letters = chr(ord("A") + remainder) + letters
        if index == 0:
            break
        index -= 1
    return letters


def _add_unique(items: list[str], value: str) -> None:
    clean = str(value or "").strip()
    if clean and clean not in items:
        items.append(clean)


def _dedupe(items: Iterable[Any]) -> list[str]:
    result: list[str] = []
    for item in items:
        _add_unique(result, str(item or ""))
    return result
