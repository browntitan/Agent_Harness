from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, List, Sequence, Tuple

from langchain_core.documents import Document

from agentic_chatbot_next.rag.status_workbooks import (
    status_metadata_for_row,
    status_metadata_for_sheet,
)


@dataclass(frozen=True)
class IndexedRow:
    row_number: int
    values: tuple[Any, ...]


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text


def _is_empty_row(values: Sequence[Any]) -> bool:
    return not any(_cell_text(item) for item in values)


def _trim_rows(rows: Sequence[IndexedRow]) -> List[IndexedRow]:
    items = [row for row in rows if not _is_empty_row(row.values)]
    return items


def _trim_trailing_empty(values: Sequence[Any]) -> List[Any]:
    items = list(values)
    while items and not _cell_text(items[-1]):
        items.pop()
    return items


def _header_score(values: Sequence[Any]) -> float:
    cells = [_cell_text(item) for item in values]
    nonempty = [item for item in cells if item]
    if len(nonempty) < 2:
        return -1.0
    unique = len(set(item.lower() for item in nonempty))
    alpha = sum(1 for item in nonempty if any(ch.isalpha() for ch in item))
    longish = sum(1 for item in nonempty if len(item) >= 3)
    return (len(nonempty) * 1.3) + (unique * 0.8) + (alpha * 0.7) + (longish * 0.4)


def _detect_header_row(rows: Sequence[IndexedRow]) -> int:
    search_window = min(len(rows), 8)
    best_index = 0
    best_score = -1.0
    for index in range(search_window):
        score = _header_score(rows[index].values)
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def _column_label(index: int) -> str:
    index = int(index)
    letters = ""
    while True:
        index, remainder = divmod(index, 26)
        letters = chr(ord("A") + remainder) + letters
        if index == 0:
            break
        index -= 1
    return letters


def _normalize_headers(values: Sequence[Any]) -> List[str]:
    headers: List[str] = []
    used: set[str] = set()
    for index, raw in enumerate(values):
        text = _cell_text(raw)
        if not text:
            text = f"Column {_column_label(index)}"
        candidate = text
        suffix = 2
        while candidate.lower() in used:
            candidate = f"{text} {suffix}"
            suffix += 1
        used.add(candidate.lower())
        headers.append(candidate)
    return headers


def _normalize_indexed_rows_with_headers(rows: Sequence[IndexedRow], headers: Sequence[str]) -> List[Tuple[int, List[str]]]:
    normalized: List[Tuple[int, List[str]]] = []
    header_len = len(headers)
    for row in rows:
        values = list(row.values)
        if len(values) < header_len:
            values.extend([None] * (header_len - len(values)))
        normalized.append((row.row_number, [_cell_text(item) for item in values[:header_len]]))
    return normalized


def _key_value_pairs(headers: Sequence[str], values: Sequence[str]) -> List[str]:
    pairs: List[str] = []
    for header, value in zip(headers, values):
        if not value:
            continue
        pairs.append(f"{header}: {value}")
    return pairs


def _sheet_range(headers: Sequence[str], rows: Sequence[IndexedRow]) -> str:
    if not rows:
        return ""
    last_col = _column_label(max(0, len(headers) - 1))
    return f"A{rows[0].row_number}:{last_col}{rows[-1].row_number}"


def _render_sheet_summary(
    workbook_name: str,
    sheet_name: str,
    headers: Sequence[str],
    data_rows: Sequence[Tuple[int, List[str]]],
    *,
    sheet_range: str,
    status_domains: Sequence[str] = (),
) -> str:
    parts = [
        f"Workbook: {workbook_name}",
        f"Sheet: {sheet_name}",
    ]
    if sheet_range:
        parts.append(f"Range: {sheet_range}")
    if headers:
        parts.append("Columns: " + ", ".join(header for header in headers if header))
    if status_domains:
        parts.append("Status Domains: " + ", ".join(str(item) for item in status_domains if str(item)))
    preview_rows: List[str] = []
    for row_number, values in data_rows[:3]:
        preview = "; ".join(_key_value_pairs(headers, values)[:6])
        if preview:
            preview_rows.append(f"Row {row_number}: {preview}")
    if preview_rows:
        parts.append("Sample Rows: " + " | ".join(preview_rows))
    return " | ".join(part for part in parts if part)


def _render_row_text(
    workbook_name: str,
    sheet_name: str,
    headers: Sequence[str],
    *,
    row_number: int,
    values: Sequence[str],
    cell_range: str,
    status_domains: Sequence[str] = (),
) -> str:
    pairs = _key_value_pairs(headers, values)
    body = "; ".join(pairs[:12]) or "No non-empty cells."
    parts = [
        f"Workbook: {workbook_name}",
        f"Sheet: {sheet_name}",
        f"Rows: {row_number}-{row_number}",
    ]
    if cell_range:
        parts.append(f"Cells: {cell_range}")
    if status_domains:
        parts.append("Status Domains: " + ", ".join(str(item) for item in status_domains if str(item)))
    parts.append(body)
    return " | ".join(part for part in parts if part)


def _build_documents_for_sheet(
    workbook_name: str,
    sheet_name: str,
    rows: Sequence[IndexedRow],
    *,
    chunk_index_start: int,
) -> List[Document]:
    trimmed = _trim_rows(rows)
    if not trimmed:
        return []

    header_index = _detect_header_row(trimmed)
    header_row = trimmed[header_index]
    headers = _normalize_headers(header_row.values)
    data_rows = _normalize_indexed_rows_with_headers(trimmed[header_index + 1 :], headers)
    sheet_status_metadata = status_metadata_for_sheet(headers)
    sheet_status_domains = [str(item) for item in (sheet_status_metadata.get("status_domains") or []) if str(item)]

    documents: List[Document] = []
    sheet_range = _sheet_range(headers, trimmed[header_index:])
    documents.append(
        Document(
            page_content=_render_sheet_summary(
                workbook_name,
                sheet_name,
                headers,
                data_rows,
                sheet_range=sheet_range,
                status_domains=sheet_status_domains,
            ),
            metadata={
                "chunk_index": chunk_index_start,
                "chunk_type": "worksheet_summary",
                "sheet_name": sheet_name,
                "row_start": int(header_row.row_number),
                "row_end": int(trimmed[-1].row_number),
                "cell_range": sheet_range,
                "is_prechunked": True,
                "status_domains": sheet_status_domains,
                "status_workbook": sheet_status_metadata,
            },
        )
    )

    next_index = chunk_index_start + 1
    max_col = _column_label(max(0, len(headers) - 1))
    for row_number, values in data_rows:
        if not any(values):
            continue
        cell_range = f"A{row_number}:{max_col}{row_number}"
        row_status_metadata = status_metadata_for_row(headers, values, row_number=row_number)
        row_status_domains = [str(item) for item in (row_status_metadata.get("status_domains") or []) if str(item)]
        documents.append(
            Document(
                page_content=_render_row_text(
                    workbook_name,
                    sheet_name,
                    headers,
                    row_number=row_number,
                    values=values,
                    cell_range=cell_range,
                    status_domains=row_status_domains,
                ),
                metadata={
                    "chunk_index": next_index,
                    "chunk_type": "spreadsheet_row",
                    "sheet_name": sheet_name,
                    "row_start": int(row_number),
                    "row_end": int(row_number),
                    "cell_range": cell_range,
                    "is_prechunked": True,
                    "status_domains": row_status_domains,
                    "status_workbook": row_status_metadata,
                },
            )
        )
        next_index += 1

    return documents


def _openpyxl_rows(path: Path) -> Iterable[Tuple[str, List[IndexedRow]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            indexed_rows = [
                IndexedRow(row_number=row_index, values=tuple(_trim_trailing_empty(cells)))
                for row_index, cells in enumerate(sheet.iter_rows(values_only=True), start=1)
            ]
            yield sheet.title, indexed_rows
    finally:
        workbook.close()


def _pandas_rows(path: Path) -> Iterable[Tuple[str, List[IndexedRow]]]:
    import pandas as pd

    workbook = pd.ExcelFile(path)
    for sheet_name in workbook.sheet_names:
        frame = workbook.parse(sheet_name=sheet_name, header=None)
        indexed_rows = [
            IndexedRow(row_number=index + 1, values=tuple(values))
            for index, values in enumerate(frame.itertuples(index=False, name=None))
        ]
        yield sheet_name, indexed_rows


def load_workbook_documents(path: Path) -> List[Document]:
    suffix = path.suffix.lower()
    sheet_iterable: Iterable[Tuple[str, List[IndexedRow]]]
    if suffix == ".xls":
        sheet_iterable = _pandas_rows(path)
    else:
        sheet_iterable = _openpyxl_rows(path)

    documents: List[Document] = []
    chunk_index = 0
    for sheet_name, rows in sheet_iterable:
        sheet_docs = _build_documents_for_sheet(
            path.name,
            sheet_name,
            rows,
            chunk_index_start=chunk_index,
        )
        documents.extend(sheet_docs)
        chunk_index += len(sheet_docs)
    return documents
